#!/usr/bin/env python3
"""Generate an Excel workbook visualising the canonical tick grid and accessibility levels.

All prices are computed via Excel formulas. The rounding step is a plain decimal
in Config!B2 (e.g. 0.000001). Changing it recalculates all tabs instantly.
"""

import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

MAX_TICK = 2600

HEADER_FILL = PatternFill(start_color="0D1B2A", end_color="0D1B2A", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)
THIN_BORDER = Border(
    left=Side(style="thin", color="CCCCCC"),
    right=Side(style="thin", color="CCCCCC"),
    top=Side(style="thin", color="CCCCCC"),
    bottom=Side(style="thin", color="CCCCCC"),
)

LEVEL_NAMES = {0: "Level 0 (÷8)", 1: "Level 1 (÷4)", 2: "Level 2 (÷2)", 3: "Level 3 (all)"}
SPACINGS = {0: 8, 1: 4, 2: 2, 3: 1}

# Row-background colors for the All Ticks tab only (light, readable).
LEVEL_ROW_FILLS = {
    0: PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid"),
    1: PatternFill(start_color="E8F0F8", end_color="E8F0F8", fill_type="solid"),
    2: PatternFill(start_color="F4F8FC", end_color="F4F8FC", fill_type="solid"),
    3: None,
}
ROW_FONT = Font(color="000000")


def finest_level(tick):
    if tick % 8 == 0:
        return 0
    if tick % 4 == 0:
        return 1
    if tick % 2 == 0:
        return 2
    return 3


def style_header(ws, col_count, row=1):
    for col in range(1, col_count + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center")
        cell.border = THIN_BORDER


def auto_width(ws, col_count, sample_rows=60):
    for col in range(1, col_count + 1):
        max_len = 0
        for row_cells in ws.iter_rows(min_col=col, max_col=col, max_row=sample_rows, values_only=False):
            for cell in row_cells:
                if cell.value is not None:
                    max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[get_column_letter(col)].width = max(max_len + 3, 12)


def write_config_sheet(wb):
    ws = wb.create_sheet(title="Config")

    ws["A1"] = "Parameter"
    ws["B1"] = "Value"
    ws["C1"] = "Description"
    style_header(ws, 3)

    ws["A2"] = "Rounding step"
    ws["B2"] = 0.000001
    ws["B2"].number_format = "0.000000000"
    ws["B2"].font = Font(color="CC0000")
    ws["B2"].border = Border(
        left=Side(style="medium", color="CC0000"),
        right=Side(style="medium", color="CC0000"),
        top=Side(style="medium", color="CC0000"),
        bottom=Side(style="medium", color="CC0000"),
    )
    ws["C2"] = "← Change this. Prices are rounded to the nearest multiple. Default 0.000001. Set to 0 for no rounding."

    ws["A3"] = "Delta (%)"
    ws["B3"] = 1.0
    ws["C3"] = "Geometric step per canonical tick, in percent."

    ws["A4"] = "MAX_TICK"
    ws["B4"] = MAX_TICK
    ws["C4"] = "Highest canonical tick index."

    ws["A5"] = "Midpoint"
    ws["B5"] = "=B4/2"
    ws["C5"] = "(derived) Half of MAX_TICK."

    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 85
    ws.freeze_panes = "A2"


def raw_price_formula(row):
    return f"1/(1+(1+Config!$B$3/100)^(Config!$B$5-A{row}))"


def rounded_price_formula(row):
    return f"IF(Config!$B$2=0,B{row},ROUND(B{row}/Config!$B$2,0)*Config!$B$2)"


def rate_formula(row, price_col="C"):
    return f'IF({price_col}{row}>0,(1/{price_col}{row}-1)*100,"∞")'


def write_all_ticks_sheet(wb):
    ws = wb.create_sheet(title="All Ticks")
    headers = [
        "Tick", "Raw Price", "Rounded Price", "Implied Rate (%)",
        "Lvl 0 (÷8)", "Lvl 1 (÷4)", "Lvl 2 (÷2)", "Lvl 3 (all)",
        "Finest Level",
    ]
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)
    style_header(ws, len(headers))

    for tick in range(0, MAX_TICK + 1):
        row = tick + 2
        lvl = finest_level(tick)

        ws.cell(row=row, column=1, value=tick)
        ws.cell(row=row, column=2, value=f"={raw_price_formula(row)}")
        ws.cell(row=row, column=2).number_format = "0.0000000000"
        ws.cell(row=row, column=3, value=f"={rounded_price_formula(row)}")
        ws.cell(row=row, column=3).number_format = "0.0000000000"
        ws.cell(row=row, column=4, value=f"={rate_formula(row)}")
        ws.cell(row=row, column=4).number_format = "0.000000"

        ws.cell(row=row, column=5, value="✓" if tick % 8 == 0 else "")
        ws.cell(row=row, column=6, value="✓" if tick % 4 == 0 else "")
        ws.cell(row=row, column=7, value="✓" if tick % 2 == 0 else "")
        ws.cell(row=row, column=8, value="✓")
        ws.cell(row=row, column=9, value=lvl)

        fill = LEVEL_ROW_FILLS[lvl]
        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=row, column=col)
            cell.font = ROW_FONT
            cell.border = THIN_BORDER
            if fill:
                cell.fill = fill
            if col >= 5:
                cell.alignment = Alignment(horizontal="center")

    ws.freeze_panes = "A2"
    auto_width(ws, len(headers))

    lc = len(headers) + 2
    ws.cell(row=1, column=lc, value="Legend").font = Font(bold=True)
    labels = ["Level 0 — every 8th tick (~8.3%)", "Level 1 — every 4th tick (~4.1%)",
              "Level 2 — every 2nd tick (~2.0%)", "Level 3 — every tick (~1.0%)"]
    for i, label in enumerate(labels):
        r = 2 + i
        ws.cell(row=r, column=lc, value=label)
        fill = LEVEL_ROW_FILLS[i]
        if fill:
            ws.cell(row=r, column=lc).fill = fill


def write_level_sheet(wb, level):
    spacing = SPACINGS[level]
    ws = wb.create_sheet(title=LEVEL_NAMES[level])
    headers = ["Tick", "Raw Price", "Rounded Price", "Implied Rate (%)", "Rate Step vs Prev (%)"]
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)
    style_header(ws, len(headers))

    row = 2
    for tick in range(0, MAX_TICK + 1, spacing):
        ws.cell(row=row, column=1, value=tick)
        ws.cell(row=row, column=2, value=f"={raw_price_formula(row)}")
        ws.cell(row=row, column=2).number_format = "0.0000000000"
        ws.cell(row=row, column=3, value=f"={rounded_price_formula(row)}")
        ws.cell(row=row, column=3).number_format = "0.0000000000"
        ws.cell(row=row, column=4, value=f"={rate_formula(row)}")
        ws.cell(row=row, column=4).number_format = "0.000000"

        if row > 2:
            ws.cell(row=row, column=5,
                    value=f'=IF(AND(D{row - 1}<>"∞",D{row}<>"∞",D{row}>0),(D{row - 1}/D{row}-1)*100,"")')
            ws.cell(row=row, column=5).number_format = "0.0000"

        for col in range(1, len(headers) + 1):
            ws.cell(row=row, column=col).border = THIN_BORDER
            ws.cell(row=row, column=col).font = ROW_FONT

        row += 1

    ws.freeze_panes = "A2"
    auto_width(ws, len(headers))


def main():
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    write_config_sheet(wb)
    write_all_ticks_sheet(wb)
    for level in range(4):
        write_level_sheet(wb, level)

    out = "tools/tick_grid.xlsx"
    wb.save(out)
    print(f"Written to {out}")
    print(f"  Config           — change B2 (rounding step) to see impact everywhere")
    print(f"  All Ticks        — {MAX_TICK + 1} rows")
    for lvl in range(4):
        count = MAX_TICK // SPACINGS[lvl] + 1
        print(f"  {LEVEL_NAMES[lvl]:15s}  — {count} ticks")


if __name__ == "__main__":
    main()
